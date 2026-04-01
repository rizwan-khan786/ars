
# # # import streamlit as st
# # # import pandas as pd
# # # import numpy as np
# # # import plotly.express as px
# # # import re
# # # from datetime import datetime

# # # st.set_page_config(page_title="GHG Summary Analyzer", layout="wide")
# # # st.title("GHG Summary Analyzer")
# # # st.caption("Upload one or more Excel files • Improved Month-wise detection")

# # # # ---------------- Helpers (kept same) ----------------
# # # def load_excel(uploaded_file):
# # #     for eng in ("openpyxl", "calamine"):
# # #         try:
# # #             return pd.ExcelFile(uploaded_file, engine=eng), eng
# # #         except Exception:
# # #             continue
# # #     raise ImportError("No Excel engine available.")

# # # def find_sheet(xls):
# # #     if "Summary Sheet" in xls.sheet_names:
# # #         return "Summary Sheet"
# # #     for s in xls.sheet_names:
# # #         if "summary" in s.lower():
# # #             return s
# # #     return xls.sheet_names[0]

# # # def header_detect_clean(df_raw: pd.DataFrame) -> pd.DataFrame:
# # #     raw = df_raw.dropna(how="all").dropna(axis=1, how="all").copy()
# # #     header_idx = raw.head(min(15, len(raw))).notna().sum(axis=1).idxmax()
# # #     df = raw.copy()
# # #     df.columns = df.loc[header_idx].astype(str).str.strip()
# # #     df = df.loc[header_idx + 1 :].reset_index(drop=True)

# # #     seen, cols = {}, []
# # #     for c in df.columns:
# # #         base = c if c and c != "nan" else "Unnamed"
# # #         seen[base] = seen.get(base, 0) + 1
# # #         cols.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
# # #     df.columns = cols

# # #     for c in df.columns:
# # #         parsed = pd.to_numeric(df[c], errors="coerce")
# # #         if parsed.notna().mean() >= 0.6:
# # #             df[c] = parsed
# # #     return df

# # # def prune_columns(df: pd.DataFrame, drop_unnamed=True, null_threshold=0.90, drop_pattern=None):
# # #     drop = []
# # #     for c in df.columns:
# # #         name = str(c).strip()
# # #         if drop_unnamed and (name.lower().startswith("unnamed") or name == "" or name.lower() == "nan"):
# # #             drop.append(c)
# # #             continue
# # #         if drop_pattern and re.match(drop_pattern, name):
# # #             drop.append(c)
# # #             continue
# # #         if df[c].isna().mean() >= null_threshold:
# # #             drop.append(c)
# # #             continue
# # #         if df[c].dtype == "object":
# # #             s = df[c].astype(str).str.strip().replace("nan", "")
# # #             if s.replace("", np.nan).dropna().empty:
# # #                 drop.append(c)
# # #                 continue
# # #     return df.drop(columns=drop), drop

# # # def arrow_safe(df: pd.DataFrame) -> pd.DataFrame:
# # #     out = df.copy()
# # #     for c in out.columns:
# # #         if out[c].dtype == "object":
# # #             out[c] = out[c].astype("string")
# # #     return out

# # # def reduce_ticks(labels, max_ticks=25):
# # #     n = len(labels)
# # #     if n <= max_ticks:
# # #         return list(range(n)), labels
# # #     step = max(1, n // max_ticks)
# # #     idxs = list(range(0, n, step))
# # #     return idxs, [labels[i] for i in idxs]

# # # def hex_to_rgba(hex_color: str, alpha: float) -> str:
# # #     try:
# # #         hc = str(hex_color).lstrip("#")
# # #         if len(hc) != 6: return hex_color
# # #         r = int(hc[0:2], 16); g = int(hc[2:4], 16); b = int(hc[4:6], 16)
# # #         return f"rgba({r},{g},{b},{alpha})"
# # #     except:
# # #         return hex_color

# # # def palette_color(palette: list, idx: int) -> str:
# # #     return palette[idx % len(palette)]

# # # PALETTES = {
# # #     "Plotly": px.colors.qualitative.Plotly,
# # #     "D3": px.colors.qualitative.D3,
# # #     "Bold": px.colors.qualitative.Bold,
# # #     "Dark24": px.colors.qualitative.Dark24,
# # #     "G10": px.colors.qualitative.G10,
# # #     "Alphabet": px.colors.qualitative.Alphabet,
# # # }

# # # # ---------------- Main App ----------------
# # # uploaded_files = st.file_uploader("Upload Excel file(s) (.xlsx)", type=["xlsx"], accept_multiple_files=True)

# # # if uploaded_files:
# # #     all_dfs = []
# # #     for uploaded in uploaded_files:
# # #         try:
# # #             xls, engine_used = load_excel(uploaded)
# # #             sheet = find_sheet(xls)
# # #             raw = pd.read_excel(uploaded, sheet_name=sheet, engine=engine_used, header=None)
# # #             df = header_detect_clean(raw)
# # #             df, _ = prune_columns(df, drop_unnamed=True, null_threshold=0.90, drop_pattern=r"^0\.0_.*")
# # #             df['Source_File'] = uploaded.name
# # #             all_dfs.append(df)
# # #         except Exception as e:
# # #             st.error(f"Error with {uploaded.name}: {e}")

# # #     if not all_dfs:
# # #         st.stop()

# # #     df_combined = pd.concat(all_dfs, ignore_index=True)

# # #     num_cols = df_combined.select_dtypes(include="number").columns.tolist()
# # #     valid_y_cols = [c for c in num_cols 
# # #                     if pd.to_numeric(df_combined[c], errors="coerce").dropna().abs().sum() != 0]

# # #     cat_cols = [c for c in df_combined.columns if c not in num_cols and c != 'Source_File']

# # #     # ---------------- SIDEBAR ----------------
# # #     with st.sidebar:
# # #         st.header("Controls")
# # #         x_col = st.selectbox("X-axis", options=cat_cols or [None])
# # #         y_cols = st.multiselect("Y-axis (numeric)", options=valid_y_cols, 
# # #                                default=valid_y_cols[:1] if valid_y_cols else [])

# # #         agg = st.radio("Aggregate by X", ["None", "sum", "mean"], index=0)
# # #         top_n = st.number_input("Top N", min_value=0, value=0, step=1)

# # #         st.markdown("### Colors")
# # #         palette_name = st.selectbox("Palette", list(PALETTES.keys()), index=0)
# # #         palette = PALETTES[palette_name]
# # #         use_custom_colors = st.checkbox("Use custom colors", value=False)

# # #         custom_color_map = {}
# # #         if use_custom_colors and y_cols:
# # #             for i, y in enumerate(y_cols):
# # #                 color_val = st.color_picker(f"{y}", value=palette_color(palette, i), key=f"col_{y}")
# # #                 custom_color_map[y] = color_val

# # #         st.markdown("**Filters**")
# # #         active_filters = {}
# # #         for c in cat_cols[:6]:
# # #             vals = df_combined[c].dropna().astype(str).unique().tolist()
# # #             if 1 < len(vals) <= 200:
# # #                 sel = st.multiselect(f"Filter {c}", options=vals, default=[], placeholder="(all)")
# # #                 if sel:
# # #                     active_filters[c] = set(sel)

# # #         chart_type = st.radio("Chart type", ["Line", "Bar", "Area"], index=1)
# # #         use_log = st.checkbox("Log scale (Y)", value=False)
# # #         rolling = st.number_input("Rolling mean window", min_value=0, value=0, step=1)
# # #         max_ticks = st.slider("Max X ticks", 5, 60, 25)

# # #     # Apply filters & aggregation (same logic)
# # #     filt = pd.Series(True, index=df_combined.index)
# # #     for col, allowed in active_filters.items():
# # #         filt &= df_combined[col].astype(str).isin(allowed)
# # #     df_f = df_combined.loc[filt].copy()

# # #     if x_col and agg != "None" and y_cols:
# # #         df_f = df_f.groupby(x_col, dropna=False)[y_cols].agg(agg).reset_index()

# # #     if top_n and y_cols and x_col:
# # #         df_f = df_f.sort_values(y_cols[0], ascending=False).head(int(top_n))

# # #     if y_cols:
# # #         y_num = df_f[y_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
# # #         df_f = df_f[y_num.abs().sum(axis=1) != 0].reset_index(drop=True)

# # #     if df_f.empty:
# # #         st.warning("No data left after filtering.")
# # #         st.stop()

# # #     st.subheader("Cleaned & Filtered Data")
# # #     st.dataframe(arrow_safe(df_f), use_container_width=True, height=400)

# # #     if not y_cols:
# # #         st.stop()

# # #     # ==================== MAIN CHARTS (unchanged) ====================
# # #     st.subheader("Main Charts")
# # #     x_vals = df_f[x_col].astype(str) if x_col else pd.Series(range(len(df_f)), name="Index")

# # #     for i, y in enumerate(y_cols):
# # #         color = custom_color_map.get(y) if use_custom_colors else palette_color(palette, i)
# # #         df_plot = pd.DataFrame({x_col or "Index": x_vals, y: pd.to_numeric(df_f[y], errors="coerce").fillna(0)})

# # #         if chart_type == "Bar":
# # #             fig = px.bar(df_plot, x=df_plot.columns[0], y=y)
# # #             fig.update_traces(marker_color=color)
# # #         elif chart_type == "Line":
# # #             fig = px.line(df_plot, x=df_plot.columns[0], y=y)
# # #             fig.update_traces(line_color=color)
# # #         else:
# # #             fig = px.area(df_plot, x=df_plot.columns[0], y=y)
# # #             fig.update_traces(line_color=color, fillcolor=hex_to_rgba(color, 0.25))

# # #         if rolling > 1:
# # #             roll = df_plot[y].rolling(rolling, min_periods=1).mean()
# # #             fig.add_scatter(x=df_plot.iloc[:,0], y=roll, mode="lines",
# # #                             name=f"Rolling", line=dict(color=hex_to_rgba(color, 0.8), dash="dash"))

# # #         idxs, labels = reduce_ticks(df_plot.iloc[:,0].tolist(), max_ticks)
# # #         fig.update_layout(
# # #             xaxis=dict(tickmode="array", tickvals=[df_plot.iloc[i,0] for i in idxs], ticktext=labels),
# # #             yaxis_type="log" if use_log else "linear",
# # #             height=480,
# # #             title=f"{y} - {chart_type}"
# # #         )
# # #         st.plotly_chart(fig, use_container_width=True)

# # #     # ==================== IMPROVED MONTH-WISE ANALYSIS ====================
# # #     st.subheader("📅 Month-wise Analysis")

# # #     # Improved detection: look for date-like column headers (2023-04-01, 2023-05, etc.)
# # #     month_col = None
# # #     date_pattern = re.compile(r'^\d{4}-\d{1,2}(-\d{1,2})?')

# # #     for col in df_f.columns:
# # #         col_str = str(col).strip()
# # #         if date_pattern.match(col_str) or any(x in col_str.lower() for x in ['month', 'period', 'date']):
# # #             month_col = col
# # #             break

# # #     if month_col:
# # #         df_month = df_f.copy()
        
# # #         # Convert column name or values to clean YYYY-MM format
# # #         try:
# # #             # If the column itself is a date string (like header "2023-04-01")
# # #             month_key = pd.to_datetime(str(month_col), errors='coerce')
# # #             if pd.notna(month_key):
# # #                 df_month['Month'] = month_key.strftime('%Y-%m')
# # #             else:
# # #                 # Try parsing values inside the column
# # #                 df_month['Month'] = pd.to_datetime(df_month[month_col], errors='coerce').dt.strftime('%Y-%m')
# # #         except:
# # #             df_month['Month'] = df_month[month_col].astype(str).str[:7]  # take first 7 chars (YYYY-MM)

# # #         df_month = df_month.dropna(subset=['Month'])
        
# # #         if not df_month.empty and y_cols:
# # #             monthly_sum = df_month.groupby('Month')[y_cols].sum().reset_index()
# # #             monthly_sum = monthly_sum.sort_values('Month')

# # #             st.success(f"✅ Using **{month_col}** for month-wise view")

# # #             for y in y_cols:
# # #                 fig_m = px.bar(monthly_sum, x='Month', y=y,
# # #                               title=f"Month-wise {y} (Total)",
# # #                               labels={'Month': 'Month'})
# # #                 fig_m.update_traces(marker_color=palette_color(palette, y_cols.index(y)))
# # #                 fig_m.update_layout(height=450)
# # #                 st.plotly_chart(fig_m, use_container_width=True)
# # #         else:
# # #             st.info("No data available for month-wise after grouping.")
# # #     else:
# # #         st.info("Still could not detect a month/date column. "
# # #                 "Please rename the column containing months (e.g. 2023-04-01) to **'Month'** and re-upload.")

# # #     # Download
# # #     st.download_button(
# # #         "Download cleaned/filtered data (CSV)",
# # #         df_f.to_csv(index=False).encode("utf-8"),
# # #         file_name="ghg_summary_cleaned.csv",
# # #         mime="text/csv"
# # #     )

# # # else:
# # #     st.info("Upload your Excel file(s) to begin.")


# # import streamlit as st
# # import pandas as pd
# # import numpy as np
# # import plotly.express as px
# # import plotly.graph_objects as go
# # from openpyxl import load_workbook
# # import datetime

# # st.set_page_config(page_title="GHG Summary Analyzer", layout="wide")
# # st.title("GHG Summary Analyzer")

# # # ─────────────────────────────────────────────────────────────────────────────
# # # CONSTANTS
# # # ─────────────────────────────────────────────────────────────────────────────
# # BASELINE_COLS     = list(range(7, 19))
# # ASSESSMENT_COLS   = list(range(20, 32))
# # BASELINE_MONTHS   = [f"2023-{m:02d}" for m in range(4,13)] + [f"2024-{m:02d}" for m in range(1,4)]
# # ASSESSMENT_MONTHS = [f"2025-{m:02d}" for m in range(4,13)] + [f"2026-{m:02d}" for m in range(1,4)]

# # GROUP_LABELS = {"A","B 1","B 2","C","C.1","C.2","C.2.1","D","D1","D2"}

# # SECTION_PARENTS = {
# #     "A1":"A","A2":"A","A11":"A","A12":"A",
# #     **{f"B1.{i}":"B1" for i in [1,2,3,4,5,6,7,8,9,10]},
# #     **{f"B2.{i}":"B2" for i in [1,2,3,4,5,6,7,8,9,10]},
# # }

# # # ─────────────────────────────────────────────────────────────────────────────
# # # PARSER
# # # ─────────────────────────────────────────────────────────────────────────────
# # def parse_excel(uploaded_file):
# #     wb = load_workbook(uploaded_file, read_only=True, data_only=True)
# #     ws = wb.active
# #     rows = list(ws.iter_rows(values_only=True))

# #     current_parent = "Other"
# #     current_group_label = "Other"   # track D1, D2, C.1, C.2 etc. for (i),(ii) disambiguation
# #     records = []

# #     for row in rows:
# #         if all(v is None for v in row):
# #             continue
# #         if any(isinstance(v, datetime.datetime) for v in row):
# #             continue

# #         raw_sec  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
# #         desc     = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
# #         freq     = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
# #         units    = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

# #         # Track parent group
# #         if raw_sec in GROUP_LABELS:
# #             current_parent = raw_sec
# #             current_group_label = raw_sec
# #             continue

# #         if not raw_sec or raw_sec.lower() in ("none","nan"):
# #             continue
# #         if desc.lower() in ("none","nan",""):
# #             continue

# #         parent = SECTION_PARENTS.get(raw_sec, current_parent)
# #         # For (i),(ii) etc. sub-rows, tag with their group for uniqueness
# #         display_sec = f"{current_group_label}>{raw_sec}" if raw_sec.startswith("(") else raw_sec

# #         for col_idx in BASELINE_COLS + ASSESSMENT_COLS:
# #             val = row[col_idx] if col_idx < len(row) else None
# #             if col_idx in BASELINE_COLS:
# #                 month_str  = BASELINE_MONTHS[col_idx - 7]
# #                 year_label = "Baseline 2023-24"
# #             else:
# #                 month_str  = ASSESSMENT_MONTHS[col_idx - 20]
# #                 year_label = "Assessment 2025-26"

# #             try:
# #                 numeric_val = float(val)
# #             except (TypeError, ValueError):
# #                 numeric_val = np.nan

# #             records.append({
# #                 "Parent":      parent,
# #                 "Section":     display_sec,
# #                 "RawSection":  raw_sec,
# #                 "Description": desc,
# #                 "Frequency":   freq,
# #                 "Units":       units,
# #                 "Year":        year_label,
# #                 "Month":       month_str,
# #                 "Value":       numeric_val,
# #             })

# #     df = pd.DataFrame(records)
# #     df["Value"] = df["Value"].fillna(0)
# #     return df


# # # ─────────────────────────────────────────────────────────────────────────────
# # # UPLOAD
# # # ─────────────────────────────────────────────────────────────────────────────
# # uploaded_files = st.file_uploader(
# #     "Upload Excel file(s) (.xlsx)", type=["xlsx"], accept_multiple_files=True
# # )
# # if not uploaded_files:
# #     st.info("Upload your Excel file(s) to begin.")
# #     st.stop()

# # all_dfs = []
# # for f in uploaded_files:
# #     try:
# #         parsed = parse_excel(f)
# #         parsed["Source_File"] = f.name
# #         all_dfs.append(parsed)
# #     except Exception as e:
# #         st.error(f"Error reading {f.name}: {e}")

# # if not all_dfs:
# #     st.stop()

# # df = pd.concat(all_dfs, ignore_index=True)
# # all_months = sorted(df["Month"].unique().tolist())

# # # Frequency/type options for dropdown (col 4 meaningful values)
# # freq_options = sorted([
# #     v for v in df["Frequency"].unique()
# #     if v and v.lower() not in ("none","nan","")
# # ])

# # # ─────────────────────────────────────────────────────────────────────────────
# # # TABS
# # # ─────────────────────────────────────────────────────────────────────────────
# # tab1, tab2 = st.tabs(["📊 Data & Charts", "🔮 Assessment Prediction"])

# # # ══════════════════════════════════════════════════════════════════════════════
# # # TAB 1
# # # ══════════════════════════════════════════════════════════════════════════════
# # with tab1:
# #     with st.sidebar:
# #         st.header("Filters")
# #         sel_freq = st.selectbox(
# #             "Type",
# #             options=["All"] + freq_options,
# #             index=0,
# #             key="t1_freq"
# #         )
# #         sel_months = st.multiselect(
# #             "Months",
# #             options=all_months,
# #             default=all_months,
# #             key="t1_months"
# #         )
# #         chart_type = st.radio("Chart type", ["Bar", "Line"], index=0, key="t1_chart")

# #     df1 = df[df["Month"].isin(sel_months)].copy()
# #     if sel_freq != "All":
# #         df1 = df1[df1["Frequency"] == sel_freq]

# #     if df1.empty:
# #         st.warning("No data for selected filters.")
# #         st.stop()

# #     # Pivot table
# #     st.subheader("Data Table — All Sections")
# #     pivot = (
# #         df1.pivot_table(
# #             index=["Parent","Section","Description","Frequency","Units","Year"],
# #             columns="Month",
# #             values="Value",
# #             aggfunc="sum"
# #         ).reset_index()
# #     )
# #     pivot.columns.name = None
# #     st.dataframe(pivot, use_container_width=True, height=440)

# #     # Charts per parent group
# #     st.subheader("Charts — All Sections")
# #     colors = px.colors.qualitative.D3

# #     for p_idx, parent in enumerate(df1["Parent"].unique().tolist()):
# #         df_p = df1[df1["Parent"] == parent]
# #         with st.expander(f"▶ {parent}", expanded=False):
# #             for s_idx, sec in enumerate(df_p["Section"].unique().tolist()):
# #                 df_sec  = df_p[df_p["Section"] == sec]
# #                 desc    = df_sec["Description"].iloc[0]
# #                 units   = df_sec["Units"].iloc[0]
# #                 title   = f"{sec} — {desc} ({units})"
# #                 df_plot = df_sec.sort_values("Month")
# #                 chart_key = f"t1_{parent}_{sec}_{p_idx}_{s_idx}"

# #                 if chart_type == "Bar":
# #                     fig = px.bar(df_plot, x="Month", y="Value", color="Year",
# #                                  title=title, barmode="group",
# #                                  color_discrete_sequence=colors)
# #                 else:
# #                     fig = px.line(df_plot, x="Month", y="Value", color="Year",
# #                                   title=title, markers=True,
# #                                   color_discrete_sequence=colors)

# #                 fig.update_layout(height=340, margin=dict(t=40,b=30),
# #                                   yaxis_title=units, xaxis_title="Month",
# #                                   legend_title="Year")
# #                 st.plotly_chart(fig, use_container_width=True, key=chart_key)

# # # ══════════════════════════════════════════════════════════════════════════════
# # # TAB 2 — Prediction
# # # ══════════════════════════════════════════════════════════════════════════════
# # with tab2:
# #     st.subheader("Assessment Year Prediction")
# #     st.caption("Uses Baseline monthly data to predict unfilled Assessment months via linear trend.")

# #     baseline_df = df[df["Year"] == "Baseline 2023-24"].copy()
# #     assess_df   = df[df["Year"] == "Assessment 2025-26"].copy()

# #     c1, c2 = st.columns([2, 4])
# #     with c1:
# #         pred_freq = st.selectbox(
# #             "Frequency / Type",
# #             options=["All"] + freq_options,
# #             index=0,
# #             key="pred_freq"
# #         )

# #     base_filtered = baseline_df if pred_freq == "All" else baseline_df[baseline_df["Frequency"] == pred_freq]

# #     valid_indicators = []
# #     for (sec, desc, units), grp in base_filtered.groupby(["Section","Description","Units"]):
# #         if grp[grp["Value"] != 0].shape[0] >= 2:
# #             valid_indicators.append(f"{sec} | {desc} ({units})")

# #     if not valid_indicators:
# #         st.info("No indicators with enough baseline data for the selected type.")
# #         st.stop()

# #     with c2:
# #         chosen = st.selectbox("Select Indicator", options=valid_indicators, key="pred_ind")

# #     sec_chosen   = chosen.split(" | ")[0]
# #     rest         = chosen.split(" | ")[1]
# #     desc_chosen  = rest.rsplit(" (", 1)[0]
# #     units_chosen = rest.rsplit("(", 1)[-1].rstrip(")")

# #     base_data = baseline_df[
# #         (baseline_df["Section"] == sec_chosen) &
# #         (baseline_df["Description"] == desc_chosen)
# #     ].sort_values("Month")[["Month","Value"]].copy()

# #     asmnt_data = assess_df[
# #         (assess_df["Section"] == sec_chosen) &
# #         (assess_df["Description"] == desc_chosen)
# #     ].sort_values("Month")[["Month","Value"]].copy()

# #     filled_months   = asmnt_data[asmnt_data["Value"] != 0]["Month"].tolist()
# #     unfilled_months = asmnt_data[asmnt_data["Value"] == 0]["Month"].tolist()

# #     base_nz = base_data[base_data["Value"] != 0].copy().reset_index(drop=True)
# #     base_nz["idx"] = np.arange(len(base_nz))

# #     if len(base_nz) >= 2:
# #         coeffs = np.polyfit(base_nz["idx"], base_nz["Value"], 1)
# #         poly   = np.poly1d(coeffs)

# #         all_assess = asmnt_data["Month"].tolist()
# #         pred_dict  = {m: max(0, poly(len(base_nz) + i)) for i, m in enumerate(all_assess)}

# #         fig = go.Figure()
# #         fig.add_trace(go.Scatter(
# #             x=base_data["Month"], y=base_data["Value"],
# #             mode="lines+markers", name="Baseline (Actual)",
# #             line=dict(color="#1f77b4", width=2), marker=dict(size=7)
# #         ))
# #         if filled_months:
# #             fd = asmnt_data[asmnt_data["Month"].isin(filled_months)]
# #             fig.add_trace(go.Scatter(
# #                 x=fd["Month"], y=fd["Value"],
# #                 mode="lines+markers", name="Assessment (Actual)",
# #                 line=dict(color="#2ca02c", width=2), marker=dict(size=7)
# #             ))
# #         if unfilled_months:
# #             fig.add_trace(go.Scatter(
# #                 x=unfilled_months,
# #                 y=[pred_dict[m] for m in unfilled_months],
# #                 mode="lines+markers", name="Predicted",
# #                 line=dict(color="#ff7f0e", width=2, dash="dash"),
# #                 marker=dict(size=8, symbol="diamond")
# #             ))

# #         fig.update_layout(
# #             title=f"Prediction: {sec_chosen} — {desc_chosen}",
# #             xaxis_title="Month", yaxis_title=units_chosen,
# #             height=420, margin=dict(t=50, b=40),
# #             legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
# #         )
# #         st.plotly_chart(fig, use_container_width=True, key="pred_main_chart")

# #         if unfilled_months:
# #             st.markdown("**Predicted values for unfilled Assessment months:**")
# #             pred_tbl = pd.DataFrame({
# #                 "Month": unfilled_months,
# #                 f"Predicted ({units_chosen})": [round(pred_dict[m], 3) for m in unfilled_months]
# #             })
# #             st.dataframe(pred_tbl, use_container_width=True, hide_index=True)
# #         else:
# #             st.success("All Assessment months already have data — no prediction needed.")

# #         direction = "↑ Increasing" if coeffs[0] > 0 else "↓ Decreasing"
# #         st.info(f"Trend: **{direction}** | Monthly change ≈ **{coeffs[0]:+.3f} {units_chosen}/month**")
# #     else:
# #         st.warning("Not enough non-zero baseline data to build a trend.")

# # # ── Download ──────────────────────────────────────────────────────────────────
# # st.download_button(
# #     "⬇ Download full data (CSV)",
# #     df.to_csv(index=False).encode("utf-8"),
# #     file_name="ghg_summary.csv",
# #     mime="text/csv"
# # )


# import streamlit as st
# import pandas as pd
# import numpy as np
# import plotly.express as px
# import plotly.graph_objects as go
# from openpyxl import load_workbook
# import datetime

# st.set_page_config(page_title="GHG Summary Analyzer", layout="wide")
# st.title("GHG Summary Analyzer")

# # ─────────────────────────────────────────────────────────────────────────────
# # CONSTANTS
# # ─────────────────────────────────────────────────────────────────────────────
# BASELINE_COLS     = list(range(7, 19))
# ASSESSMENT_COLS   = list(range(20, 32))
# BASELINE_MONTHS   = [f"2023-{m:02d}" for m in range(4,13)] + [f"2024-{m:02d}" for m in range(1,4)]
# ASSESSMENT_MONTHS = [f"2025-{m:02d}" for m in range(4,13)] + [f"2026-{m:02d}" for m in range(1,4)]

# GROUP_LABELS = {"A","B 1","B 2","C","C.1","C.2","C.2.1","D","D1","D2"}

# SECTION_PARENTS = {
#     "A1":"A","A2":"A","A11":"A","A12":"A",
#     **{f"B1.{i}":"B1" for i in [1,2,3,4,5,6,7,8,9,10]},
#     **{f"B2.{i}":"B2" for i in [1,2,3,4,5,6,7,8,9,10]},
# }

# # ─────────────────────────────────────────────────────────────────────────────
# # PARSER
# # ─────────────────────────────────────────────────────────────────────────────
# def parse_excel(uploaded_file):
#     wb = load_workbook(uploaded_file, read_only=True, data_only=True)
#     ws = wb.active
#     rows = list(ws.iter_rows(values_only=True))

#     current_parent = "Other"
#     current_group_label = "Other"   # track D1, D2, C.1, C.2 etc. for (i),(ii) disambiguation
#     records = []

#     for row in rows:
#         if all(v is None for v in row):
#             continue
#         if any(isinstance(v, datetime.datetime) for v in row):
#             continue

#         raw_sec  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
#         desc     = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
#         freq     = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
#         units    = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

#         # Track parent group
#         if raw_sec in GROUP_LABELS:
#             current_parent = raw_sec
#             current_group_label = raw_sec
#             continue

#         if not raw_sec or raw_sec.lower() in ("none","nan"):
#             continue
#         if desc.lower() in ("none","nan",""):
#             continue

#         parent = SECTION_PARENTS.get(raw_sec, current_parent)
#         # For (i),(ii) etc. sub-rows, tag with their group for uniqueness
#         display_sec = f"{current_group_label}>{raw_sec}" if raw_sec.startswith("(") else raw_sec

#         for col_idx in BASELINE_COLS + ASSESSMENT_COLS:
#             val = row[col_idx] if col_idx < len(row) else None
#             if col_idx in BASELINE_COLS:
#                 month_str  = BASELINE_MONTHS[col_idx - 7]
#                 year_label = "Baseline 2023-24"
#             else:
#                 month_str  = ASSESSMENT_MONTHS[col_idx - 20]
#                 year_label = "Assessment 2025-26"

#             try:
#                 numeric_val = float(val)
#             except (TypeError, ValueError):
#                 numeric_val = np.nan

#             records.append({
#                 "Parent":      parent,
#                 "Section":     display_sec,
#                 "RawSection":  raw_sec,
#                 "Description": desc,
#                 "Frequency":   freq,
#                 "Units":       units,
#                 "Year":        year_label,
#                 "Month":       month_str,
#                 "Value":       numeric_val,
#             })

#     df = pd.DataFrame(records)
#     df["Value"] = df["Value"].fillna(0)
#     return df


# # ─────────────────────────────────────────────────────────────────────────────
# # UPLOAD
# # ─────────────────────────────────────────────────────────────────────────────
# uploaded_files = st.file_uploader(
#     "Upload Excel file(s) (.xlsx)", type=["xlsx"], accept_multiple_files=True
# )
# if not uploaded_files:
#     st.info("Upload your Excel file(s) to begin.")
#     st.stop()

# all_dfs = []
# for f in uploaded_files:
#     try:
#         parsed = parse_excel(f)
#         parsed["Source_File"] = f.name
#         all_dfs.append(parsed)
#     except Exception as e:
#         st.error(f"Error reading {f.name}: {e}")

# if not all_dfs:
#     st.stop()

# df = pd.concat(all_dfs, ignore_index=True)
# all_months = sorted(df["Month"].unique().tolist())

# # Frequency/type options for dropdown (col 4 meaningful values)
# freq_options = sorted([
#     v for v in df["Frequency"].unique()
#     if v and v.lower() not in ("none","nan","")
# ])

# # ─────────────────────────────────────────────────────────────────────────────
# # TABS
# # ─────────────────────────────────────────────────────────────────────────────
# tab1, tab2 = st.tabs(["📊 Data & Charts", "🔮 Assessment Prediction"])

# freq_display_options = []

# for (sec, desc, units), grp in df.groupby(["Section", "Description", "Units"]):
#     freq_display_options.append((sec, desc, units))
# # ══════════════════════════════════════════════════════════════════════════════
# # TAB 1
# # ══════════════════════════════════════════════════════════════════════════════
# with tab1:
#     with st.sidebar:
#         st.header("Filters")
#         sel_freq = st.selectbox(
#             "Type",
#             options=["All"] + freq_options,
#             index=0,
#             key="t1_freq"
#         )
#         sel_months = st.multiselect(
#             "Months",
#             options=all_months,
#             default=all_months,
#             key="t1_months"
#         )
#         chart_type = st.radio("Chart type", ["Bar", "Line"], index=0, key="t1_chart")

#     df1 = df[df["Month"].isin(sel_months)].copy()
#     if sel_freq != "All":
#         df1 = df1[df1["Frequency"] == sel_freq]

#     if df1.empty:
#         st.warning("No data for selected filters.")
#         st.stop()

#     # Pivot table
#     st.subheader("Data Table — All Sections")
#     pivot = (
#         df1.pivot_table(
#             index=["Parent","Section","Description","Frequency","Units","Year"],
#             columns="Month",
#             values="Value",
#             aggfunc="sum"
#         ).reset_index()
#     )
#     pivot.columns.name = None
#     st.dataframe(pivot, use_container_width=True, height=440)

#     # Charts per parent group
#     st.subheader("Charts — All Sections")
#     colors = px.colors.qualitative.D3

#     for p_idx, parent in enumerate(df1["Parent"].unique().tolist()):
#         df_p = df1[df1["Parent"] == parent]
#         with st.expander(f"▶ {parent}", expanded=False):
#             for s_idx, sec in enumerate(df_p["Section"].unique().tolist()):
#                 df_sec  = df_p[df_p["Section"] == sec]
#                 desc    = df_sec["Description"].iloc[0]
#                 units   = df_sec["Units"].iloc[0]
#                 title   = f"{sec} — {desc} ({units})"
#                 df_plot = df_sec.sort_values("Month")
#                 chart_key = f"t1_{parent}_{sec}_{p_idx}_{s_idx}"

#                 if chart_type == "Bar":
#                     fig = px.bar(df_plot, x="Month", y="Value", color="Year",
#                                  title=title, barmode="group",
#                                  color_discrete_sequence=colors)
#                 else:
#                     fig = px.line(df_plot, x="Month", y="Value", color="Year",
#                                   title=title, markers=True,
#                                   color_discrete_sequence=colors)

#                 fig.update_layout(height=340, margin=dict(t=40,b=30),
#                                   yaxis_title=units, xaxis_title="Month",
#                                   legend_title="Year")
#                 st.plotly_chart(fig, use_container_width=True, key=chart_key)

# # ══════════════════════════════════════════════════════════════════════════════
# # TAB 2 — Prediction
# # ══════════════════════════════════════════════════════════════════════════════
# with tab2:
#     st.subheader("Assessment Year Prediction")
#     st.caption("Uses Baseline monthly data to predict unfilled Assessment months via linear trend.")

#     baseline_df = df[df["Year"] == "Baseline 2023-24"].copy()
#     assess_df   = df[df["Year"] == "Assessment 2025-26"].copy()

#     # c1, c2 = st.columns([2, 4])
#     # with c1:
#     #     pred_freq = st.selectbox(
#     #         "Frequency / Type",
#     #         options=["All"] + freq_options,
#     #         index=0,
#     #         key="pred_freq"
#     #     )

#     # base_filtered = baseline_df if pred_freq == "All" else baseline_df[baseline_df["Frequency"] == pred_freq]
#     base_filtered = baseline_df.copy()

#     valid_indicators = []
#     for (sec, desc, units), grp in base_filtered.groupby(["Section","Description","Units"]):
#         if grp[grp["Value"] != 0].shape[0] >= 2:
#             valid_indicators.append(f"{sec} | {desc} ({units})")

#     if not valid_indicators:
#         st.info("No indicators with enough baseline data for the selected type.")
#         st.stop()

#     # with c2:
#     #     chosen = st.selectbox("Select Indicator", options=valid_indicators, key="pred_ind")
#     chosen = st.selectbox("Select Indicator", options=valid_indicators, key="pred_ind")

#     sec_chosen   = chosen.split(" | ")[0]
#     rest         = chosen.split(" | ")[1]
#     desc_chosen  = rest.rsplit(" (", 1)[0]
#     units_chosen = rest.rsplit("(", 1)[-1].rstrip(")")

#     base_data = baseline_df[
#         (baseline_df["Section"] == sec_chosen) &
#         (baseline_df["Description"] == desc_chosen)
#     ].sort_values("Month")[["Month","Value"]].copy()

#     asmnt_data = assess_df[
#         (assess_df["Section"] == sec_chosen) &
#         (assess_df["Description"] == desc_chosen)
#     ].sort_values("Month")[["Month","Value"]].copy()

#     filled_months   = asmnt_data[asmnt_data["Value"] != 0]["Month"].tolist()
#     unfilled_months = asmnt_data[asmnt_data["Value"] == 0]["Month"].tolist()

#     base_nz = base_data[base_data["Value"] != 0].copy().reset_index(drop=True)
#     base_nz["idx"] = np.arange(len(base_nz))

#     if len(base_nz) >= 2:
#         coeffs = np.polyfit(base_nz["idx"], base_nz["Value"], 1)
#         poly   = np.poly1d(coeffs)

#         all_assess = asmnt_data["Month"].tolist()
#         pred_dict  = {m: max(0, poly(len(base_nz) + i)) for i, m in enumerate(all_assess)}

#         # Build unified bar chart: all months, colour-coded by type
#         rows_chart = []
#         for _, r in base_data.iterrows():
#             rows_chart.append({"Month": r["Month"], "Value": r["Value"], "Type": "Baseline"})
#         for _, r in asmnt_data.iterrows():
#             if r["Month"] in filled_months:
#                 rows_chart.append({"Month": r["Month"], "Value": r["Value"], "Type": "Assessment (Actual)"})
#             else:
#                 rows_chart.append({"Month": r["Month"], "Value": pred_dict[r["Month"]], "Type": "Predicted"})

#         df_chart = pd.DataFrame(rows_chart)
#         color_map = {
#             "Baseline":            "#1f77b4",
#             "Assessment (Actual)": "#2ca02c",
#             "Predicted":           "#ff7f0e",
#         }

#         fig = px.bar(
#             df_chart, x="Month", y="Value", color="Type",
#             color_discrete_map=color_map,
#             title=f"{sec_chosen} — {desc_chosen}",
#             barmode="group",
#         )
#         for trace in fig.data:
#             if trace.name == "Predicted":
#                 trace.marker.opacity = 0.75

#         fig.update_layout(
#             height=420, margin=dict(t=50, b=40),
#             yaxis_title=units_chosen, xaxis_title="Month",
#             legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
#         )
#         st.plotly_chart(fig, use_container_width=True, key="pred_main_chart")

#         if unfilled_months:
#             st.markdown("**Predicted values for unfilled Assessment months:**")
#             pred_tbl = pd.DataFrame({
#                 "Month": unfilled_months,
#                 f"Predicted ({units_chosen})": [round(pred_dict[m], 3) for m in unfilled_months]
#             })
#             st.dataframe(pred_tbl, use_container_width=True, hide_index=True)
#         else:
#             st.success("All Assessment months already have data — no prediction needed.")

#         direction    = "Up (Increasing)" if coeffs[0] > 0 else "Down (Decreasing)"
#         slope        = round(coeffs[0], 4)
#         intercept    = round(coeffs[1], 4)
#         avg_baseline = round(float(base_nz["Value"].mean()), 3)
#         n_points     = len(base_nz)

#         with st.expander("How was this prediction calculated?", expanded=True):
#             st.markdown(f"""
# **Method: Linear Regression (Least Squares)**

# The prediction uses the {n_points} non-zero months from the **Baseline year (2023-24)**
# to fit a straight-line trend, then extends that line into the **Assessment year (2025-26)**.

# ---

# **Step-by-step:**

# 1. **Collect baseline data** — took all {n_points} months where the value was not zero.
#    Average baseline value: **{avg_baseline} {units_chosen}**

# 2. **Assign a number to each month** — each month gets an index (0, 1, 2 ... {n_points-1})
#    so the formula can work on plain numbers instead of dates.

# 3. **Fit a straight line** using the formula:
#    `Value = slope x month_index + intercept`
#    Slope = **{slope}** | Intercept = **{intercept}**

# 4. **Extend the line forward** — Assessment months continue the index
#    from {n_points} up to {n_points+11}, giving the predicted values.

# 5. **Floor at zero** — any predicted value below 0 is clamped to 0.

# ---

# **Trend direction: {direction}**
# Each month the value changes by approximately **{slope:+.3f} {units_chosen}**.

# *Note: This is a simple linear extrapolation based on the available baseline months.
# It works well for stable metrics. It does not account for seasonality or sudden
# operational changes.*
# """)
#     else:
#         st.warning("Not enough non-zero baseline data to build a trend.")

# # ── Download ──────────────────────────────────────────────────────────────────
# st.download_button(
#     "⬇ Download full data (CSV)",
#     df.to_csv(index=False).encode("utf-8"),
#     file_name="ghg_summary.csv",
#     mime="text/csv"
# )


import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
import datetime

st.set_page_config(page_title="GHG Summary Analyzer", layout="wide")
st.title("GHG Summary Analyzer")

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
BASELINE_COLS     = list(range(7, 19))
ASSESSMENT_COLS   = list(range(20, 32))
BASELINE_MONTHS   = [f"2023-{m:02d}" for m in range(4,13)] + [f"2024-{m:02d}" for m in range(1,4)]
ASSESSMENT_MONTHS = [f"2025-{m:02d}" for m in range(4,13)] + [f"2026-{m:02d}" for m in range(1,4)]

GROUP_LABELS = {"A","B 1","B 2","C","C.1","C.2","C.2.1","D","D1","D2"}

SECTION_PARENTS = {
    "A1":"A","A2":"A","A11":"A","A12":"A",
    **{f"B1.{i}":"B1" for i in [1,2,3,4,5,6,7,8,9,10]},
    **{f"B2.{i}":"B2" for i in [1,2,3,4,5,6,7,8,9,10]},
}

# ─────────────────────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────────────────────
def parse_excel(uploaded_file):
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    current_parent = "Other"
    current_group_label = "Other"
    records = []

    for row in rows:
        if all(v is None for v in row):
            continue
        if any(isinstance(v, datetime.datetime) for v in row):
            continue

        raw_sec  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        desc     = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        freq     = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        units    = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

        if raw_sec in GROUP_LABELS:
            current_parent = raw_sec
            current_group_label = raw_sec
            continue

        if not raw_sec or raw_sec.lower() in ("none","nan"):
            continue
        if desc.lower() in ("none","nan",""):
            continue

        parent = SECTION_PARENTS.get(raw_sec, current_parent)
        display_sec = f"{current_group_label}>{raw_sec}" if raw_sec.startswith("(") else raw_sec

        for col_idx in BASELINE_COLS + ASSESSMENT_COLS:
            val = row[col_idx] if col_idx < len(row) else None
            if col_idx in BASELINE_COLS:
                month_str  = BASELINE_MONTHS[col_idx - 7]
                year_label = "Baseline 2023-24"
            else:
                month_str  = ASSESSMENT_MONTHS[col_idx - 20]
                year_label = "Assessment 2025-26"

            try:
                numeric_val = float(val)
            except (TypeError, ValueError):
                numeric_val = np.nan

            records.append({
                "Parent":      parent,
                "Section":     display_sec,
                "RawSection":  raw_sec,
                "Description": desc,
                "Frequency":   freq,
                "Units":       units,
                "Year":        year_label,
                "Month":       month_str,
                "Value":       numeric_val,
            })

    df = pd.DataFrame(records)
    df["Value"] = df["Value"].fillna(0)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# UPLOAD
# ─────────────────────────────────────────────────────────────────────────────
uploaded_files = st.file_uploader(
    "Upload Excel file(s) (.xlsx)", type=["xlsx"], accept_multiple_files=True
)
if not uploaded_files:
    st.info("Upload your Excel file(s) to begin.")
    st.stop()

all_dfs = []
for f in uploaded_files:
    try:
        parsed = parse_excel(f)
        parsed["Source_File"] = f.name
        all_dfs.append(parsed)
    except Exception as e:
        st.error(f"Error reading {f.name}: {e}")

if not all_dfs:
    st.stop()

df = pd.concat(all_dfs, ignore_index=True)
all_months = sorted(df["Month"].unique().tolist())

# Frequency/type options
freq_options = sorted([
    v for v in df["Frequency"].unique()
    if v and v.lower() not in ("none","nan","")
])

# Units options
units_options = sorted([
    v for v in df["Units"].unique()
    if v and v.lower() not in ("none","nan","")
])

# Build section → description mapping for sidebar display
sec_desc_map = (
    df[["Section","Description"]]
    .drop_duplicates("Section")
    .set_index("Section")["Description"]
    .to_dict()
)

# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📊 Data & Charts", "🔮 Assessment Prediction"])

freq_display_options = []

for (sec, desc, units), grp in df.groupby(["Section", "Description", "Units"]):
    freq_display_options.append((sec, desc, units))

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    with st.sidebar:
        st.header("Filters")

        # ── Type filter ──────────────────────────────────────────────────────
        # sel_freq = st.selectbox(
        #     "Type",
        #     options=["All"] + freq_options,
        #     index=0,
        #     key="t1_freq"
        # )

        # ── Units filter (NEW) ───────────────────────────────────────────────
        sel_units = st.selectbox(
            "Units",
            options=["All"] + units_options,
            index=0,
            key="t1_units"
        )

        # ── Months filter ────────────────────────────────────────────────────
        sel_months = st.multiselect(
            "Months",
            options=all_months,
            default=all_months,
            key="t1_months"
        )

        # ── Chart type ───────────────────────────────────────────────────────
        chart_type = st.radio("Chart type", ["Bar", "Line"], index=0, key="t1_chart")

        # ── Column C — Section Names ─────────────────────────────────────────
        st.markdown("---")
        st.subheader("📋 Section Names (Column C)")

        # Get unique sections with their descriptions, sorted
        section_info = (
            df[["Section","Description","Units"]]
            .drop_duplicates(subset=["Section"])
            .sort_values("Section")
        )

        # Apply current type / units filter so the list stays in sync
        filtered_section_info = section_info.copy()
        # if sel_freq != "All":
        #     valid_secs = df[df["Frequency"] == sel_freq]["Section"].unique()
        #     filtered_section_info = filtered_section_info[
        #         filtered_section_info["Section"].isin(valid_secs)
        #     ]
        if sel_units != "All":
            valid_secs2 = df[df["Units"] == sel_units]["Section"].unique()
            filtered_section_info = filtered_section_info[
                filtered_section_info["Section"].isin(valid_secs2)
            ]

        for _, row_s in filtered_section_info.iterrows():
            sec_label = row_s["Section"]
            desc_label = row_s["Description"]
            unit_label = row_s["Units"] if row_s["Units"] and row_s["Units"].lower() not in ("none","nan","") else "—"
            st.markdown(
                f"<div style='margin-bottom:4px;'>"
                f"<span style='font-weight:600;color:#1565C0;'>{sec_label}</span>"
                f"<br><span style='font-size:0.78em;color:#444;'>{desc_label}</span>"
                f"<br><span style='font-size:0.72em;color:#888;'>Unit: {unit_label}</span>"
                f"</div>",
                unsafe_allow_html=True
            )

    # ── Apply filters ─────────────────────────────────────────────────────────
    df1 = df[df["Month"].isin(sel_months)].copy()
    # if sel_freq != "All":
    #     df1 = df1[df1["Frequency"] == sel_freq]
    if sel_units != "All":
        df1 = df1[df1["Units"] == sel_units]

    if df1.empty:
        st.warning("No data for selected filters.")
        st.stop()

    # Pivot table
    st.subheader("Data Table — All Sections")
    pivot = (
        df1.pivot_table(
            index=["Parent","Section","Description","Frequency","Units","Year"],
            columns="Month",
            values="Value",
            aggfunc="sum"
        ).reset_index()
    )
    pivot.columns.name = None
    st.dataframe(pivot, use_container_width=True, height=440)

    # Charts per parent group
    st.subheader("Charts — All Sections")
    colors = px.colors.qualitative.D3

    for p_idx, parent in enumerate(df1["Parent"].unique().tolist()):
        df_p = df1[df1["Parent"] == parent]
        with st.expander(f"▶ {parent}", expanded=False):
            for s_idx, sec in enumerate(df_p["Section"].unique().tolist()):
                df_sec  = df_p[df_p["Section"] == sec]
                desc    = df_sec["Description"].iloc[0]
                units   = df_sec["Units"].iloc[0]
                title   = f"{sec} — {desc} ({units})"
                df_plot = df_sec.sort_values("Month")
                chart_key = f"t1_{parent}_{sec}_{p_idx}_{s_idx}"

                if chart_type == "Bar":
                    fig = px.bar(df_plot, x="Month", y="Value", color="Year",
                                 title=title, barmode="group",
                                 color_discrete_sequence=colors)
                else:
                    fig = px.line(df_plot, x="Month", y="Value", color="Year",
                                  title=title, markers=True,
                                  color_discrete_sequence=colors)

                fig.update_layout(height=340, margin=dict(t=40,b=30),
                                  yaxis_title=units, xaxis_title="Month",
                                  legend_title="Year")
                st.plotly_chart(fig, use_container_width=True, key=chart_key)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Prediction
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("Assessment Year Prediction")
    st.caption("Uses Baseline monthly data to predict unfilled Assessment months via linear trend.")

    baseline_df = df[df["Year"] == "Baseline 2023-24"].copy()
    assess_df   = df[df["Year"] == "Assessment 2025-26"].copy()

    base_filtered = baseline_df.copy()

    valid_indicators = []
    for (sec, desc, units), grp in base_filtered.groupby(["Section","Description","Units"]):
        if grp[grp["Value"] != 0].shape[0] >= 2:
            valid_indicators.append(f"{sec} | {desc} ({units})")

    if not valid_indicators:
        st.info("No indicators with enough baseline data for the selected type.")
        st.stop()

    chosen = st.selectbox("Select Indicator", options=valid_indicators, key="pred_ind")

    sec_chosen   = chosen.split(" | ")[0]
    rest         = chosen.split(" | ")[1]
    desc_chosen  = rest.rsplit(" (", 1)[0]
    units_chosen = rest.rsplit("(", 1)[-1].rstrip(")")

    base_data = baseline_df[
        (baseline_df["Section"] == sec_chosen) &
        (baseline_df["Description"] == desc_chosen)
    ].sort_values("Month")[["Month","Value"]].copy()

    asmnt_data = assess_df[
        (assess_df["Section"] == sec_chosen) &
        (assess_df["Description"] == desc_chosen)
    ].sort_values("Month")[["Month","Value"]].copy()

    filled_months   = asmnt_data[asmnt_data["Value"] != 0]["Month"].tolist()
    unfilled_months = asmnt_data[asmnt_data["Value"] == 0]["Month"].tolist()

    base_nz = base_data[base_data["Value"] != 0].copy().reset_index(drop=True)
    base_nz["idx"] = np.arange(len(base_nz))

    if len(base_nz) >= 2:
        coeffs = np.polyfit(base_nz["idx"], base_nz["Value"], 1)
        poly   = np.poly1d(coeffs)

        all_assess = asmnt_data["Month"].tolist()
        pred_dict  = {m: max(0, poly(len(base_nz) + i)) for i, m in enumerate(all_assess)}

        rows_chart = []
        for _, r in base_data.iterrows():
            rows_chart.append({"Month": r["Month"], "Value": r["Value"], "Type": "Baseline"})
        for _, r in asmnt_data.iterrows():
            if r["Month"] in filled_months:
                rows_chart.append({"Month": r["Month"], "Value": r["Value"], "Type": "Assessment (Actual)"})
            else:
                rows_chart.append({"Month": r["Month"], "Value": pred_dict[r["Month"]], "Type": "Predicted"})

        df_chart = pd.DataFrame(rows_chart)
        color_map = {
            "Baseline":            "#1f77b4",
            "Assessment (Actual)": "#2ca02c",
            "Predicted":           "#ff7f0e",
        }

        fig = px.bar(
            df_chart, x="Month", y="Value", color="Type",
            color_discrete_map=color_map,
            title=f"{sec_chosen} — {desc_chosen}",
            barmode="group",
        )
        for trace in fig.data:
            if trace.name == "Predicted":
                trace.marker.opacity = 0.75

        fig.update_layout(
            height=420, margin=dict(t=50, b=40),
            yaxis_title=units_chosen, xaxis_title="Month",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        st.plotly_chart(fig, use_container_width=True, key="pred_main_chart")

        if unfilled_months:
            st.markdown("**Predicted values for unfilled Assessment months:**")
            pred_tbl = pd.DataFrame({
                "Month": unfilled_months,
                f"Predicted ({units_chosen})": [round(pred_dict[m], 3) for m in unfilled_months]
            })
            st.dataframe(pred_tbl, use_container_width=True, hide_index=True)
        else:
            st.success("All Assessment months already have data — no prediction needed.")

        direction    = "Up (Increasing)" if coeffs[0] > 0 else "Down (Decreasing)"
        slope        = round(coeffs[0], 4)
        intercept    = round(coeffs[1], 4)
        avg_baseline = round(float(base_nz["Value"].mean()), 3)
        n_points     = len(base_nz)

        with st.expander("How was this prediction calculated?", expanded=True):
            st.markdown(f"""
**Method: Linear Regression (Least Squares)**

The prediction uses the {n_points} non-zero months from the **Baseline year (2023-24)**
to fit a straight-line trend, then extends that line into the **Assessment year (2025-26)**.

---

**Step-by-step:**

1. **Collect baseline data** — took all {n_points} months where the value was not zero.
   Average baseline value: **{avg_baseline} {units_chosen}**

2. **Assign a number to each month** — each month gets an index (0, 1, 2 ... {n_points-1})
   so the formula can work on plain numbers instead of dates.

3. **Fit a straight line** using the formula:
   `Value = slope x month_index + intercept`
   Slope = **{slope}** | Intercept = **{intercept}**

4. **Extend the line forward** — Assessment months continue the index
   from {n_points} up to {n_points+11}, giving the predicted values.

5. **Floor at zero** — any predicted value below 0 is clamped to 0.

---

**Trend direction: {direction}**
Each month the value changes by approximately **{slope:+.3f} {units_chosen}**.

*Note: This is a simple linear extrapolation based on the available baseline months.
It works well for stable metrics. It does not account for seasonality or sudden
operational changes.*
""")
    else:
        st.warning("Not enough non-zero baseline data to build a trend.")

# ── Download ──────────────────────────────────────────────────────────────────
st.download_button(
    "⬇ Download full data (CSV)",
    df.to_csv(index=False).encode("utf-8"),
    file_name="ghg_summary.csv",
    mime="text/csv"
)